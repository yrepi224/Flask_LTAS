# -*- coding:utf-8 -*-

from module import upload_db, select_raw_data, upload_sorted_data, all_select_data, share_select_data
from openpyxl import Workbook, load_workbook
from pathlib import Path


def upload_csv(filename):
    report_name = select_raw_data(filename)
    raw_data = dict()
    # 파일 오픈
    excel_name = Path(report_name)
    file_name = excel_name.stem
    excel_data = load_workbook(excel_name, data_only=True)
    # raw_data 데이터 밀어넣기
    for i in excel_data.sheetnames:
        app_sheet = excel_data[i]
        name_app = i
        for row in app_sheet.rows:
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            raw_data[f'{name_app}__{row_value[3]}__{row_value[1]}'] = raw_data.get(
                row_value[0], [row_value[0], row_value[1], name_app, row_value[3], file_name])
    # DB 업로드
    upload_sorted_data(raw_data)
