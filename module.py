# -*- coding:utf-8 -*-
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from datetime import datetime
import glob
import os
import csv
bad_host = ['facebook', 'google.co.kr', 'google.com', 'amazonaws', 'ubuntu', 'canonical', 'googlemail', 'akamaitechnologies.com', '1e100.net', 'service.game-mode.net', 'localhost', 'DESKTOP-18JBMRB',
            'cloudfront.net', '1e100.net', 'display.ad.g.daum.net', 'googleusercontent.com', 'doubleclick', 'ec2', 'd1iskralo6mo11.cloudfront.net', 'beacons.gvt2.com', 'samsungiotcloud',
            'measurement', 'android', '.local', 'akamai', 'gvt1.com', 'apple', 'onesignal', 'gstatic', 'mcafee', 'googleapis', 'MyoungehsiPhone', 'Domain']


# 데이터베이스 insert
def upload_db(raw_data):
    conn = psycopg2.connect(database="ryu",
                            user="sungwonryu",
                            host="127.0.0.1",
                            password="7887",
                            port="5432")
    cur = conn.cursor()
    cnt = 0
    fcnt = 0
    for key, value in raw_data.items():
        try:
            cur.execute(
                f"""INSERT INTO public."RawData"("Pkey", "host", "ip", "app", "service", "filename", "date")VALUES('{key}', '{value[0]}', '{value[1]}', '{value[2]}', '{value[3]}', '{value[4]}', '{datetime.today().strftime("%Y-%m-%d")}')""")
            conn.commit()
            cnt = cnt + 1
            print('successfully imported data!         '+str(cnt)+'  '+key)
        except:
            fcnt = fcnt + 1
            print('fail count ('+str(fcnt)+')')
            conn.rollback()
    print('Imported ('+str(cnt)+') lines of All Data!')
    print('Failed ('+str(fcnt)+') lines of All Data!')
    print('------------------------------------------------------------------------------')


# 데이터베이스 all_data
def all_select_data(report):
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True, size=11)
    bd = Side(style='thick', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    connection = psycopg2.connect(database="ryu",
                                  user="sungwonryu",
                                  host="127.0.0.1",
                                  password="7887",
                                  port="5432")

    try:
        with connection.cursor() as cursor:
            query = f'''SELECT "host", "ip", "app", "service", "duplicate", "date" FROM public."AllData" WHERE "filename" LIKE'%{report}%'ORDER BY "app", "service", "ip"'''
            cursor.execute(query)
            rs = cursor.fetchall()
            wb = Workbook()
            # DB 모든 데이터 엑셀로
            name = 'NULL'
            num = 0
            for row in rs:
                if name != row[2]:
                    wb.create_sheet(index=num, title=row[2])
                    ws = wb.get_sheet_by_name(row[2])
                    name = row[2]
                    num = num + 1
                    ws.append(['Domain', 'IP', 'App', 'Service', '공통여부', 'Date'])
                ws.append(row)
                if row[4] is True:
                    ws['A'+str(ws.max_row)].style = highlight
            wb.remove(wb['Sheet'])
            wb.save(report+' 계열 보고서.xlsx')
            print(f'Successfully Saved {report}.xlsx')
    finally:
        connection.close()
        wb.close()

# 데이터베이스 Share_data


def share_select_data(report):
    connection = psycopg2.connect(database="ryu",
                                  user="sungwonryu",
                                  host="127.0.0.1",
                                  password="7887",
                                  port="5432")
    try:
        with connection.cursor() as cursor:
            query = f'''SELECT "host", "ip", "app", "service", "date" FROM public."ShareData" WHERE "filename" LIKE'%{report}%'ORDER BY "ip", "service", "app"'''
            cursor.execute(query)
            rs = cursor.fetchall()
            wb = Workbook()
            ws = wb.active
            wb['Sheet'].title = '공통'
            # 첫행 입력
            ws.append(('Domain', 'IP', 'App', 'Service', 'Date'))

            # DB 모든 데이터 엑셀로
            for row in rs:
                ws.append(row)
            wb.save(report+' 공통.xlsx')
            print(f'Successfully Saved {report} 공통.xlsx')
    finally:
        connection.close()
        wb.close()


# 데이터베이스 raw_select
def select_raw_data(filename):

    report = filename
    connection = psycopg2.connect(database="ryu",
                                  user="sungwonryu",
                                  host="127.0.0.1",
                                  password="7887",
                                  port="5432")

    try:
        with connection.cursor() as cursor:
            query = f'''SELECT "host", "ip", "app", "service", "date" FROM public."RawData" WHERE "filename" LIKE'%{report}%'ORDER BY "app", "service"'''
            cursor.execute(query)
            rs = cursor.fetchall()
            wb = Workbook()
            # DB 모든 데이터 엑셀로
            name = 'NULL'
            num = 0
            for row in rs:
                if name != row[2]:
                    wb.create_sheet(index=num, title=row[2])
                    ws = wb.get_sheet_by_name(row[2])
                    name = row[2]
                    num = num + 1
                ws.append(row)
            wb.remove(wb['Sheet'])
            wb.save(report+'.xlsx')
            report_name = (f'{report}.xlsx')
            return report_name
    finally:
        connection.close()
        wb.close()


def upload_sorted_data(raw_data):
    conn = psycopg2.connect(database="ryu",
                            user="sungwonryu",
                            host="127.0.0.1",
                            password="7887",
                            port="5432")

    cur = conn.cursor()
    host_data = dict()
    # host_data로 데이터 이전 및 호스트명 통일
    for key, value in raw_data.items():
        host_data[value[2]+'__'+value[3]+'__'+value[1]] = [value[0],
                                                           value[1], value[2], value[3], value[4], 'FALSE']
    for key1, host_val in host_data.items():
        for key, raw_val in raw_data.items():
            if host_val[1] == raw_val[1] and host_val[0] == host_val[1] and raw_val[0] != raw_val[1]:
                host_data[key1] = [raw_val[0], host_val[1],
                                   host_val[2], host_val[3], host_val[4], 'FALSE']
                break

    # 공통 호스트 변경 작업 1
    for key1, value1 in host_data.items():
        for key, value2 in host_data.items():
            if value1[1] == value2[1] and value1[2] != value2[2] or value1[0] == value2[0] and value1[2] != value2[2]:
                host_data[key1] = [value1[0], value1[1],
                                   value1[2], value1[3], value1[4], 'TRUE']

    # 공통 호스트 변경 작업 2
    for key1, value1 in host_data.items():
        for value2 in host_data.items():
            if value1[5] == 'FALSE':
                if value1[0] == value2[0] and value2[5] == 'TRUE':
                    host_data[key1] = [value1[0], value1[1],
                                       value1[2], value1[3], value1[4], 'TRUE']

    # 데이터베이스 insert - 전체 데이터
    cnt = 0
    fcnt = 0
    for key, value in host_data.items():
        no_save = True
        for host in bad_host:
            if host in value[0]:
                no_save = False
                break
        if no_save is True:
            try:
                cur.execute(f"""INSERT INTO public."AllData"("Pkey", "host", "ip", "app", "service", "filename", "duplicate", "date")
                VALUES('{key}', '{value[0]}', '{value[1]}', '{value[2]}', '{value[3]}', '{value[4]}', '{value[5]}', '{datetime.today().strftime("%Y-%m-%d")}')""")
                conn.commit()
                cnt = cnt + 1
                print(f'successfully imported data!         {str(cnt)}  {key}')
            except:
                fcnt = fcnt + 1
                print('fail count ('+str(fcnt)+')')
                conn.rollback()
    print(f'Imported {str(cnt)} lines of All Data!')
    print(f'Failed {str(fcnt)} lines of All Data!')
    print('------------------------------------------------------------------------------')

    # 데이터베이스 insert - 공통데이터
    cnt = 0
    fcnt = 0
    for key, value in host_data.items():
        if value[5] == 'TRUE':
            no_save = True
            for host in bad_host:
                if host in value[0]:
                    no_save = False
                    break
            if no_save is True:
                try:
                    cur.execute(f"""INSERT INTO public."ShareData"("Pkey", "host", "ip", "app", "service", "filename", "date")
                    VALUES('{key}', '{value[0]}', '{value[1]}', '{value[2]}', '{value[3]}', '{value[4]}', '{datetime.today().strftime("%Y-%m-%d")}')""")
                    conn.commit()
                    cnt = cnt + 1
                    print(f'successfully imported data!         {str(cnt)}  {key}')
                except:
                    fcnt = fcnt + 1
                    print(f'fail count {str(fcnt)}')
                    conn.rollback()
    conn.close()
    print(f'Imported {str(cnt)} lines of Share Data!')
    print(f'Failed {str(fcnt)} lines of Share Data!')


def emergency_upload(foldername, filename):
    raw_data = dict()
    print('Try it Multiple Times')
    name_app = foldername
    input_path = f'/Users/sungwonryu/Documents/GitHub/Flask_LTAS/csv/{name_app}'
    for input_file in glob.glob(os.path.join(input_path, '*.csv')):
        with open(input_file, 'r', newline='') as csv_in_file:
            filereader = csv.reader(csv_in_file)
            # 전체 데이터 딕셔너리 host_data[0] = Domain, [1] = IP, [2] = App, [3] = Service, [4] = Filename, [5] = 공통여부
            cnt = 0
            for row_value in filereader:
                raw_data[row_value[2]+'__'+row_value[3]+'__'+row_value[1]] = raw_data.get(
                    row_value[0], [row_value[0], row_value[1], row_value[2], row_value[3], filename])

    conn = psycopg2.connect(database="ryu",
                            user="sungwonryu",
                            host="127.0.0.1",
                            password="7887",
                            port="5432")
    cur = conn.cursor()
    cnt = 0
    fcnt = 0
    for key, value in raw_data.items():
        try:
            cur.execute(f'''INSERT INTO public."RawData"("Pkey", "host", "ip", "app", "service", "filename")VALUES('{key}', '{value[0]}', '{value[1]}', '{value[2]}', '{value[3]}', '{value[4]}')''')
            conn.commit()
            cnt = cnt + 1
            print('successfully imported data!         '+str(cnt)+'  '+key)
        except:
            fcnt = fcnt + 1
            print('fail count ('+str(fcnt)+')')
            conn.rollback()
    print('Imported ('+str(cnt)+') lines of All Data!')
    print('Failed ('+str(fcnt)+') lines of All Data!')
    print('------------------------------------------------------------------------------')

# **개별실행**
# emergency_upload()
# upload_db()
# all_select_data()
# share_select_data()
# select_raw_data()
# upload_sorted_data(raw_data필요)
